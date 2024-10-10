def extract_tables_from_html(file_content, progress_bar, total_zones, status_text):
    from lxml import etree
    from io import BytesIO, StringIO
    import pandas as pd
    import streamlit as st
    from funcs import process_dataframe_for_styler

    # Process the file content and extract tables
    zone_tables = {}  # Dictionary to store tables for each zone
    context = etree.iterparse(BytesIO(file_content), html=True, events=('start', 'end'))
    zone_name = None
    tables = []
    collecting = False
    table_title = ''
    table_count = 0
    processed_zones = 0

    for event, elem in context:
        if event == 'start':
            if elem.tag == 'p' and 'For:' in ''.join(elem.itertext()) and 'Entire Facility' not in ''.join(elem.itertext()):
                zone_name_elem = elem.find('b')
                if zone_name_elem.text is not None:
                    zone_name = zone_name_elem.text.strip()
                else:
                    zone_name = 'Unknown Zone'
                collecting = True
                status_text.write(f"Processing zone: {zone_name}")
                tables = []
                table_count = 0
            elif collecting and elem.tag == 'b':
                table_title = ''.join(elem.itertext()).strip()
            elif collecting and elem.tag == 'table':
                table_html = etree.tostring(elem, encoding='unicode')
                try:
                    # Update pd.read_html to skip the first row, use the second row as headers,
                    # and set the first column as index
                    df = pd.read_html(
                        StringIO(table_html),
                        flavor='lxml',
                        header=0,
                        index_col=0
                    )[0]

                    df = process_dataframe_for_styler(df)

                    tables.append((table_title, df))
                except Exception as e:
                    print(f"Failed to read table '{table_title}' for zone '{zone_name}': {e}")
                table_count += 1
                if table_count == 6:
                    collecting = False
                    # Store the tables for the zone
                    zone_tables[zone_name] = tables
                    tables = []
                    table_count = 0
                    # Update the progress bar
                    processed_zones += 1
                    if total_zones > 0:
                        progress = processed_zones / total_zones
                        progress_bar.progress(progress)
        elem.clear()
    del context
    return zone_tables

def count_zones_in_toc(file_content):
    from io import StringIO
    import re

    # Read the file content line by line up to a reasonable limit
    max_lines = 2000  # Adjust as needed based on file structure
    toc_started = False
    toc_content = ''
    zone_links = []
    
    # Decode the file content for line-by-line processing
    content_stream = StringIO(file_content.decode('utf-8', errors='ignore'))

    for _ in range(max_lines):
        line = content_stream.readline()
        if not line:
            break  # End of file reached
        if '<p><b>Zone Component Load Summary' in line:
            toc_started = True
        elif toc_started and '<p><b>' in line and 'Zone Component Load Summary' not in line:
            break  # End of Table of Contents
        if toc_started:
            toc_content += line

    # Use regex to find all zone links in the TOC
    zone_links = re.findall(r'<a href="#ZoneComponentLoadSummary::(.*?)">', toc_content)
    return len(zone_links)

def apply_table_format(ws, start_row, start_col, df, title, is_cooling=True):
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Define the fill colors for cooling and heating tables
    cooling_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    heating_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
    
    # Apply the appropriate fill color based on the type of table (cooling or heating)
    title_fill = cooling_fill if is_cooling else heating_fill
    
    # Write the title in the first cell, apply the fill color, bold, and alignment
    title_mid_col = start_col + df.shape[1] // 2
    for col_num in range(start_col, start_col + df.shape[1] + 1):  # Apply fill to all title cells
        cell = ws.cell(row=start_row, column=col_num)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_num == title_mid_col:
            cell.value = title  # Place the title in the middle cell
            cell.font = Font(bold=True)

    # Add the DataFrame headers (use 'Metric' instead of 'Index')
    for col_num, header in enumerate(['Metric'] + list(df.columns), start=start_col):
        cell = ws.cell(row=start_row + 1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

    # Add the DataFrame data and ensure numbers are written as numeric types
    for row_num, row_data in enumerate(df.itertuples(index=True), start=start_row + 2):
        for col_num, value in enumerate(row_data, start=start_col):
            # Convert values that are numeric (but stored as strings) to float or int
            try:
                value = float(value) if '.' in str(value) else int(value)
            except (ValueError, TypeError):
                pass  # Keep as string if it can't be converted
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

    # Return the last row where the table was inserted
    return start_row + len(df) + 2

def generate_excel(zone_tables):
    from io import BytesIO
    import openpyxl 
    from openpyxl.styles import Border, Side

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zone Tables"

    # Setting the thin border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Loop through the zones and add cooling and heating tables side by side
    current_row = 1
    for zone_name, tables in zone_tables.items():
        cooling_tables = [(title, df) for title, df in tables if 'Cooling' in title]
        heating_tables = [(title, df) for title, df in tables if 'Heating' in title]
        
        max_table_length = max(len(cooling_tables), len(heating_tables))
        
        # For each table in the cooling and heating, lay them out side by side
        for i in range(max_table_length):
            # Cooling table
            if i < len(cooling_tables):
                cooling_title, cooling_df = cooling_tables[i]
                apply_table_format(ws, current_row, 1, cooling_df, f"{zone_name} - {cooling_title}", is_cooling=True)
            # Heating table (starting from column 'K' which is 11th column)
            if i < len(heating_tables):
                heating_title, heating_df = heating_tables[i]
                apply_table_format(ws, current_row, 11, heating_df, f"{zone_name} - {heating_title}", is_cooling=False)

            # Move to the next row after placing the side-by-side tables
            current_row += max(cooling_df.shape[0], heating_df.shape[0]) + 4

    # Save the workbook into a BytesIO object
    wb.save(output)
    output.seek(0)  # Set pointer to the beginning of the file
    return output
    
def clean_filename(name):
    import re
    return re.sub(r'[\\/*?:"<>|]', "_", name)

def process_dataframe_for_styler(df):
    import pandas as pd
    import re

    """
    Convert applicable columns to floats, round them, and convert back to strings with two decimal places.
    
    Parameters:
    - df (pd.DataFrame): The original DataFrame.
    
    Returns:
    - pd.DataFrame: The processed DataFrame with formatted strings.
    """
    # Identify numeric columns based on dtypes
    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
    
    # Additionally, identify columns with numeric data stored as objects
    for col in df.columns:
        if df[col].dtype == 'object':
            # Attempt to detect if the column is numeric by checking the first non-null value
            sample = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            if sample is not None and re.match(r'^-?\d+(\.\d+)?$', str(sample)):
                numeric_cols.append(col)
    
    # Remove duplicates
    numeric_cols = list(set(numeric_cols))
    
    # Process each numeric column
    for col in numeric_cols:
        # Remove non-numeric characters if any
        df[col] = df[col].astype(str).apply(lambda x: re.sub(r'[^\d\.\-]', '', x))
        
        # Convert to float, coercing errors to NaN
        df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Round to two decimal places
        df[col] = df[col].round(2)
        
        # Convert back to string with two decimal places
        df[col] = df[col].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")
    
    return df